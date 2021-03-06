import pandas as pd
from openpyxl import load_workbook
from string import Template
import numpy as np
from csv import reader
import xlrd
import os
import datetime


locationDictionary = {'Astation' : 'Aurora Station Store', 'A PATP' : 'Aurora Station Store', 'ASTATION PATP' : 'Aurora Station Store', 'ASTATION' : 'Aurora Station Store', 'Main Office' : 'Admin', 'MAIN OFFICE' : "Admin", 'TIRE CENTER' : 'Aurora Tire Center', 'AW PATP' : 'Aurora West A Stop', 'Awest PATP' : 'Aurora West A Stop', 'AWPATP' : 'Aurora West A Stop', 'AWEST PATP' : 'Aurora West A Stop', 'CLAY CENTER' : 'Clay Center', 'Dannbrog PATP' : 'Dannebrog Station', 'DANNEBORG' : 'Dannebrog Station', 'Dannbrog Station' : 'Dannebrog Station', 'DANNEROG' : 'Dannebrog Station',  'Elwood' : 'Elwood Station', 'ELWOOD' : 'Elwood Station', 'GIBBON': 'Gibbon', 'GI PATP' : 'Grand Island', 'GISLAND' : 'Grand Island', 'GISLAND PATP' : 'Grand Island', 'Gisland' : 'Grand Island', 'GI FEED MILL' : 'Grand Island Grain & Feed', "GRANT" : 'Grant', 'GRANT PATP' : 'Grant', 'HARDY' : 'Hardy', "Harvard PATP" : "Harvard", "HARVARD" : "Harvard", "Hastings PATP" : 'Hastings', "HASTINGS"  : 'Hastings', "KEARNEY": 'Kearney', "Keen PATP" : "Keene", 'KEEN PATP' : "Keene", 'MINDEN' : 'Minden', 'MINDEN PATP': "Minden", 'Minden' : 'Minden', "POCOMOKE" : 'Pocomoke', 'StPaul' : 'St Paul Station', 'ST PAUL'  : 'St Paul Station', 'ST PAUL PATP' : 'St Paul Station', 'SUPERIOR' : 'Superior', 'UPLAND' : 'Upland', 'UPLAND PATP' : 'Upland', 'YORK PATP' : 'York', 'YORK' : 'York', 'York' : 'York'}

#readFile1 = r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\JuneExport.xlsx'
readFile1 = input ("Please input the path to the file. This file path should only be a file that contains one month of PostDates for CC transactions.\n ")


wb = load_workbook(readFile1, data_only=True)
exportSheet = wb['Export']

with open(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\WriteFile1.txt', 'w') as f1:
    f1.write("TRCNumber" + "\t" +	"AccountNumber" + "\t" + "AccountType" + "\t" + "AccountName"  + "\t" +	"PostDate" + "\t" + "Reference" + "\t" + "AdditionalReference" + "\t" + "Amount" + "\t" +	"Description" + "\t" + "Type" + "\t" + "Text"  + "\t" +	"Type"  + "\t" + "LocNum"  + "\t" +	"LocName" + "\n")
with open(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\WriteFile1.txt', 'a') as f1:    
    for row in range(2, exportSheet.max_row + 1):
        TRC_NUMBER = exportSheet['A' + str(row)].value
        ACCOUNT_NUM = exportSheet['B' + str(row)].value
        ACCOUNT_TYPE = exportSheet['C' + str(row)].value
        ACCOUNT_NAME = exportSheet['D' + str(row)].value
        POST_DATE = exportSheet['E' + str(row)].value
        REFERENCE = exportSheet['F' + str(row)].value
        ADDITIONAL_REFERENCE = exportSheet['G' + str(row)].value
        AMOUNT = exportSheet['H' + str(row)].value
        DESCRIPTION = exportSheet['I' + str(row)].value
        TYPE_1 = exportSheet['J' + str(row)].value
        TEXT = exportSheet['K' + str(row)].value
        TYPE_2 = exportSheet['L' + str(row)].value
        LOC_NUM = exportSheet['M' + str(row)].value
        LOC_NAME = exportSheet['N' + str(row)].value
        for x, y in locationDictionary.items():
            if LOC_NAME == x:
                LOC_NAME = y
                break
            else:
                LOC_NAME = LOC_NAME
        f1.write(str(TRC_NUMBER) + "\t" + str(ACCOUNT_NUM) + "\t" + str(ACCOUNT_TYPE) + "\t" + str(ACCOUNT_NAME) + "\t" + str(POST_DATE) + "\t" + str(REFERENCE) + "\t" + str(ADDITIONAL_REFERENCE) + "\t" + str(AMOUNT) + "\t" + str(DESCRIPTION) + "\t" + str(TYPE_1) + "\t" + str(TEXT) + "\t" + str(TYPE_2) + "\t" + str(LOC_NUM) + "\t" + str(LOC_NAME) + '\n')


print("WriteFile1.txt is done!")
wb.close()
print("Closing workbook.")


#readFile2 = r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\Jun Credit Card Reconciliation to GL.xlsx'
readFile2 = input ("Please input the path to the file. This file path should contain transactions spreading over several months. This should be your GL transaction history file.\n ")

wb2 = load_workbook(readFile2, data_only=True)
glLedger = wb2['GeneralLedgerDetailReportList']
startOnRowNumber = input("Enter the number of Excel row that you would like to start on. The row that you will want to start on will be 7 days prior to the start of the month's transactions in the first file that you entered in the previous step. \n (This must be in integer format.) \n\n")
startOnRowNumber = int(startOnRowNumber.strip())

with open(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\WriteFile2.txt', 'w') as f2:
    f2.write("GL" + "\t" +	"PC" + "\t" + "SourceName" + "\t" + 'Account' + "\t" +	'BlankColumn' + "\t" +	'Name' + "\t" +	'CM' + "\t" + 'Loc'  + "\t" + 'Date' + "\t" + 'Ticket' + "\t" +	'Type' + "\t" +	'Debit' + "\t" + 'Credit' + "\t" + 'Qty' + "\t" + 'RunningBalance' + "\t" + 'SourceDescription' + "\t" + 'LocationName' + "\t" + 'ImportCleared'  + "\t" + 'Comments' + '\n')
with open(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\WriteFile2.txt', 'a') as f2:    
    for row in range(startOnRowNumber, glLedger.max_row + 1):
        GL = glLedger['A' + str(row)].value
        PC = glLedger['B' + str(row)].value
        SOURCE_NAME = glLedger['C' + str(row)].value
        ACCOUNT = glLedger['D' + str(row)].value
        BLANK = glLedger['E' + str(row)].value
        NAME = glLedger['F' + str(row)].value
        CM = glLedger['G' + str(row)].value
        LOC = glLedger['H' + str(row)].value
        DATE = glLedger['I' + str(row)].value
        TICKET = glLedger['J' + str(row)].value
        TYPE = glLedger['K' + str(row)].value
        DEBIT = glLedger['L' + str(row)].value
        CREDIT = glLedger['M' + str(row)].value
        QTY = glLedger['N' + str(row)].value
        RUNNING_BALANCE = glLedger['O' + str(row)].value
        SOURCE_DESCRIPTION = glLedger['P' + str(row)].value
        LOCATION_NAME = glLedger['Q' + str(row)].value
        IMPORT_CLEARED = glLedger['R' + str(row)].value
        COMMENTS = glLedger['S' + str(row)].value
        f2.write(str(GL) + "\t" + str(PC) + "\t" + str(SOURCE_NAME) + "\t" + str(ACCOUNT) + "\t" + str(BLANK) + "\t" + str(NAME) + "\t" + str(CM) + "\t" + str(LOC) + "\t" + str(DATE) + "\t" + str(TICKET) + "\t" + str(TYPE) + "\t" + str(DEBIT) + "\t" + str(CREDIT) + "\t" + str(QTY) + "\t" + str(RUNNING_BALANCE) + "\t" + str(SOURCE_DESCRIPTION) + "\t" + str(LOCATION_NAME) + "\t" + str(IMPORT_CLEARED) + "\t" + str(COMMENTS) + '\n')

print("WriteFile2.txt is done!")
wb.close()
print("Closing workbook.")

df1 = pd.read_csv(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\WriteFile1.txt', engine="python", sep='\t')
df1 = df1.sort_values(by=['LocName', 'PostDate', 'Amount'], ascending=False)



df2 = pd.read_csv(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\WriteFile2.txt', engine="python", sep='\t')
df2 = df2.sort_values(by=["LocationName", "Date", "Debit"], ascending=False)

df2 = df2[df2.Type != 'clear']
df2 = df2[df2.Type != 'clera']
df2 = df2[df2.Type != 'ctear']

df3 = pd.merge(df1, df2, how="inner", left_on=['LocName', 'Amount'], right_on=['LocationName', 'Debit'])

print(df3)

#print(df3.info())

df3['DateDifference'] = df3['PostDate'].astype('datetime64') - df3['Date'].astype('datetime64')

df3['DateDifference'] = df3['DateDifference'].dt.days

df3.loc[(df3['DateDifference'] <=7) & (df3['DateDifference'] >=0), 'InRange?'] = 'True'
df3.loc[(df3['DateDifference'] >7) | (df3['DateDifference'] <0), 'InRange?'] = 'False'

#df3.info()

#print(df3)

#df3.info()

inRange = df3[df3['InRange?'] == 'True']

notInRange = df3[df3['InRange?'] == 'False']

df3.sort_values(by=["PostDate", "LocName"])

print("***************IN RANGE************************")
print(inRange)
print("***************NOT IN RANGE************************")
print(notInRange)

df3.to_csv(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\WriteFile3.txt', sep='\t')

print("WriteFile3 is done. Check this file for matching transactions")