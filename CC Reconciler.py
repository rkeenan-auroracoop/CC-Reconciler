import os
import pprint
import openpyxl
from openpyxl import load_workbook
import pandas as pd


READFILE1 = (r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\Copy of 01-05 Banking Entries.xlsx')
wb1 = load_workbook(READFILE1, data_only=True)
SHEET1 = wb1['Sheet1']

for row in range(3, SHEET1.max_row + 1):
    with open(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\WriteFile1.txt', 'a') as f1:
        POSTDATE = SHEET1['A' + str(row)].value
        if str(POSTDATE) <= "2020-01-07 00:00:00": 
            REFERENCENUM = SHEET1['B' + str(row)].value
            ADDITIONALREF = SHEET1['C' + str(row)].value
            AMOUNT = SHEET1['D' + str(row)].value
            STATUS = SHEET1['E' + str(row)].value
            DESCRIPTION = SHEET1['F' + str(row)].value
            TYPE = SHEET1['G' + str(row)].value
            TEXT = SHEET1['H' + str(row)].value
            CC_COMPANY = SHEET1['I' + str(row)].value
            LOCATION_NUM = SHEET1['J' + str(row)].value
            LOCATION_NAME = SHEET1['K' + str(row)].value
            print(str(POSTDATE) + '\t' + str(REFERENCENUM) + '\t' + str(ADDITIONALREF) + '\t' + str(AMOUNT) + '\t' + str(STATUS) + '\t' + str(DESCRIPTION) + '\t' + str(TYPE) + '\t' + str(TEXT) + '\t' + str(CC_COMPANY) + '\t' + str(LOCATION_NUM) + '\t' + str(LOCATION_NAME) + '\n')
            f1.write(str(POSTDATE) + '\t' + str(REFERENCENUM) + '\t' + str(ADDITIONALREF) + '\t' + str(AMOUNT) + '\t' + str(STATUS) + '\t' + str(DESCRIPTION) + '\t' + str(TYPE) + '\t' + str(TEXT) + '\t' + str(CC_COMPANY) + '\t' + str(LOCATION_NUM) + '\t' + str(LOCATION_NAME) + '\n')
        else:
            break

print("Your write file #1 is done!")
wb1.close()


READFILE2 = (r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\Copy of May Credit Card Reconciliation to GL.xlsx')
wb2 = load_workbook(READFILE2, data_only=True)
SHEET2 = wb2['GeneralLedgerDetailReportList']
DATE_RANGE = pd.to_datetime('2020-01-01').date()

for row in range(7, SHEET1.max_row + 1):
    with open(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\WriteFile2.txt', 'a') as f1:
        GL = SHEET2['A' + str(row)].value
        PC = SHEET2['B' + str(row)].value
        SOURCE_NAME = SHEET2['C' + str(row)].value
        ACCOUNT = SHEET2['D' + str(row)].value
        PERSON = SHEET2['E' + str(row)].value
        NAME = SHEET2['I' + str(row)].value       
        DATE = SHEET2['I' + str(row)].value
        DATE = pd.to_datetime(DATE).date()
        TICKET = SHEET2['J' + str(row)].value
        TYPE = SHEET2['K' + str(row)].value
        DEBIT = SHEET2['L' + str(row)].value
        CREDIT = SHEET2['M' + str(row)].value
        QUANTITY = SHEET2['N' + str(row)].value
        RUNNING_BALANCE = SHEET2['O' + str(row)].value
        SOURCE_DESCRIPTION = SHEET2['P' + str(row)].value
        ACE_LOCATION = SHEET2['Q' + str(row)].value
        IMPORT_CLEARED = SHEET2['R' + str(row)].value
        COMMENTS = SHEET2['S' + str(row)].value

        if DATE < DATE_RANGE:
            print(str(DATE) + " is in range")
            f1.write(str(GL) + "\t" + str(PC) + "\t" + str(SOURCE_NAME) + "\t" + str(ACCOUNT) + "\t" + str(PERSON) + "\t" + str(NAME) + "\t" + str(DATE) + "\t" + str(TICKET) + "\t" + str(TYPE) + "\t" + str(DEBIT) + "\t" + str(CREDIT) + "\t" + str(QUANTITY) + "\t" + str(RUNNING_BALANCE) + "\t" + str(SOURCE_DESCRIPTION) + "\t" + str(ACE_LOCATION) + "\t" + str(IMPORT_CLEARED) + "\t" + str(COMMENTS) + '\n')
        else:
            break

wb2.close()