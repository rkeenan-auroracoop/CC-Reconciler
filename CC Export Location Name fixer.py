import pandas as pd
from openpyxl import load_workbook
from string import Template

wb = load_workbook(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\Development\Account Reconciler\JuneExport.xlsx', data_only=True)

Admin = ('Main Office', 'Admin', 'MAIN OFFICE')
AuroraStoreStation = ('Astation', 'A PATP', 'ASTATION PATP', 'ASTATION')
AuroraTireCenter = ('TIRE CENTER')
AuroraWestAStop = ('AW PATP', 'Awest PATP', 'AWPATP', 'AWEST PATP')
ClayCenter = ('CLAY CENTER')
Dannebrog = ('Dannbrog PATP', 'DANNEBORG', 'Dannbrog Station', 'DANNEROG')
Elwood = ('Elwood', 'ELWOOD')
Gibbon = ('GIBBON', 'GIBBON')
GrandIslandGrainAndFeed = ('GI FEED MILL')
GrandIsland = ('GI PATP', 'GISLAND', 'GISLAND PATP', 'Gisland')
Grant = ("Grant", 'GRANT PATP')
Hardy = ("HARDY")
Harvard = ("Harvard PATP", "HARVARD" )
Hastings = ("Hastings PATP", "HASTINGS")
Kearney = ("KEARNEY",)
Keene = ("Keen PATP", 'KEEN PATP')
Minden = ('MINDEN', 'MINDEN PATP', 'Minden')
Pocomoke = ("POCOMOKE")
StPaul = ('StPaul', 'ST PAUL', 'ST PAUL PATP')
Superior = ('SUPERIOR')
Upland = ('UPLAND', 'UPLAND PATP')
York = ('YORK PATP', 'YORK', 'York')

def fixLocationNames():
    exportSheet = wb.active
    for row in range(2, exportSheet.max_row + 1):
        LOCATIONNAME = exportSheet['N' + str(row)].value
        if LOCATIONNAME in Admin:
            print("This is a match!")
            #do find and replace based on selection in conditional tree
        elif LOCATIONNAME in AuroraStoreStation:
            print("This is a match!")
        elif LOCATIONNAME in AuroraTireCenter:
            print("This is a match!")
        elif LOCATIONNAME in AuroraWestAStop:
            print("This is a match!")
        elif LOCATIONNAME in ClayCenter:
            print("This is a match!")
        elif LOCATIONNAME in Dannebrog:
            print("This is a match!")
        elif LOCATIONNAME in Elwood:
            print("This is a match!")
        elif LOCATIONNAME in Gibbon:
            print("This is a match!")
        elif LOCATIONNAME in GrandIslandGrainAndFeed:
            print("This is a match!")
        elif LOCATIONNAME in GrandIsland:
            print("This is a match!")
        elif LOCATIONNAME in Grant:
            print("This is a match!")
        elif LOCATIONNAME in Hardy:
            print("This is a match!")
        elif LOCATIONNAME in Harvard:
            print("This is a match!")
        elif LOCATIONNAME in Hastings:
            print("This is a match!")
        elif LOCATIONNAME in Kearney:
            print("This is a match!")
        elif LOCATIONNAME in Keene:
            print("This is a match!")
        elif LOCATIONNAME in Minden:
            print("This is a match!")
        elif LOCATIONNAME in Pocomoke:
            print("This is a match!")
        elif LOCATIONNAME in StPaul:
            print("This is a match!")
        elif LOCATIONNAME in Superior:
            print("This is a match!")
        elif LOCATIONNAME in Upland:
            print("This is a match!")
        elif LOCATIONNAME in York:
            print("This is a match!")
        else:
            print("Not a match!" + "\t" +LOCATIONNAME)

fixLocationNames()